"""
NeatSheet Automation Script
----------------------------
Reads 'messy_data.csv', cleans Name/Phone/Email data,
removes duplicate emails, and exports a formatted Excel file.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import sys

# -----------------------------
# STEP 1: Load the source file
# -----------------------------
SOURCE_FILE = "messy_data.csv"
OUTPUT_FILE = "NeatSheet_Output.xlsx"

# Check the file actually exists before trying to read it.
# This avoids a confusing crash and gives a clear error instead.
if not os.path.exists(SOURCE_FILE):
    print(f"ERROR: Could not find '{SOURCE_FILE}' in this folder.")
    print("Make sure messy_data.csv is in the same directory as this script.")
    sys.exit(1)

print(f"Loading data from '{SOURCE_FILE}'...")
df = pd.read_csv(SOURCE_FILE)

# -----------------------------
# STEP 2: Clean the 'Name' column
# -----------------------------
# .str.strip() removes leading/trailing whitespace.
# .str.title() converts to Title Case (e.g. "john SMITH" -> "John Smith")
if "Name" in df.columns:
    df["Name"] = df["Name"].astype(str).str.strip().str.title()
else:
    print("WARNING: No 'Name' column found - skipping name cleanup.")

# -----------------------------
# STEP 3: Clean the 'Phone' column
# -----------------------------
# We use a regex to strip out anything that ISN'T a digit (0-9).
# This removes spaces, hyphens, parentheses, dots, etc.
if "Phone" in df.columns:
    df["Phone"] = (
        df["Phone"]
        .astype(str)
        .str.replace(r"[^0-9]", "", regex=True)
    )
else:
    print("WARNING: No 'Phone' column found - skipping phone cleanup.")

# -----------------------------
# STEP 4: Remove duplicate rows based on Email
# -----------------------------
# We try to auto-detect the email column name in case it's
# capitalized differently (Email, email, E-mail, etc.)
email_col = None
for col in df.columns:
    if col.strip().lower() in ("email", "e-mail", "email address"):
        email_col = col
        break

if email_col:
    before_count = len(df)
    # Normalize email casing/whitespace first so "Bob@Mail.com "
    # and "bob@mail.com" are correctly recognized as duplicates.
    df[email_col] = df[email_col].astype(str).str.strip().str.lower()
    df = df.drop_duplicates(subset=email_col, keep="first")
    after_count = len(df)
    print(f"Removed {before_count - after_count} duplicate row(s) based on '{email_col}'.")
else:
    print("WARNING: No email column found - skipping duplicate removal.")

# -----------------------------
# STEP 5: Save to a nicely formatted Excel file
# -----------------------------
print(f"Saving cleaned data to '{OUTPUT_FILE}'...")

# Write the dataframe to Excel using openpyxl as the engine
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Clean Data")

    # Grab the worksheet object so we can style it
    worksheet = writer.sheets["Clean Data"]

    # --- Style the header row ---
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:  # row 1 = header row
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # --- Auto-fit column widths based on content ---
    for i, column in enumerate(df.columns, start=1):
        max_length = max(
            df[column].astype(str).map(len).max(),  # longest data value
            len(str(column))                          # header length
        )
        col_letter = get_column_letter(i)
        worksheet.column_dimensions[col_letter].width = max_length + 4

    # --- Freeze header row so it stays visible when scrolling ---
    worksheet.freeze_panes = "A2"

print(f"Done! '{OUTPUT_FILE}' has been created successfully.")